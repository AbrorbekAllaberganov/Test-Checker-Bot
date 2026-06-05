"""
app/services/excel.py — Natijalarni Excel (.xlsx) formatida eksport qilish servisi.
"""
from __future__ import annotations

import io
from datetime import datetime
from typing import Optional

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.models.attempt import Attempt
from app.models.student import Student
from app.models.test import Test
from app.models.titul import Titul
from app.models.group import Group


def _style_sheet(ws) -> None:
    """Excel varag'ini chiroyli va qulay qilish uchun stillash."""
    # Sarlavha font va fill
    header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
    
    # Oddiy kataklar uchun
    data_font = Font(name="Calibri", size=11)
    thin_side = Side(border_style="thin", color="D3D3D3")
    thin_border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    
    # Markazlashtirish va tekislash
    center_align = Alignment(horizontal="center", vertical="center")
    left_align = Alignment(horizontal="left", vertical="center")
    
    # Birinchi qatorni (header) stillash
    for col in range(1, ws.max_column + 1):
        cell = ws.cell(row=1, column=col)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = center_align
        cell.border = thin_border

    # Ma'lumotlarni stillash
    for row in range(2, ws.max_row + 1):
        for col in range(1, ws.max_column + 1):
            cell = ws.cell(row=row, column=col)
            cell.font = data_font
            cell.border = thin_border
            
            # Formatga ko'ra tekislash
            val = cell.value
            if isinstance(val, (int, float)):
                cell.alignment = center_align
                if col == ws.max_column - 2 and isinstance(val, float):  # Foiz ustuni
                    cell.number_format = '0.0"%"'
            elif isinstance(val, datetime):
                cell.alignment = center_align
                cell.value = val.strftime("%d.%m.%Y %H:%M")
            else:
                # Text ustuni
                if str(val) in ["Ha", "Yo'q"]:
                    cell.alignment = center_align
                else:
                    cell.alignment = left_align

    # Qator balandligi
    ws.row_dimensions[1].height = 25
    for r in range(2, ws.max_row + 1):
        ws.row_dimensions[r].height = 20

    # Ustun kengliklarini avtomatik to'g'rilash
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            val = cell.value
            if val is not None:
                if isinstance(val, datetime):
                    val_str = val.strftime("%d.%m.%Y %H:%M")
                elif isinstance(val, float):
                    val_str = f"{val:.1f}%"
                else:
                    val_str = str(val)
                max_len = max(max_len, len(val_str))
        ws.column_dimensions[col_letter].width = max(max_len + 4, 12)


def _to_bytes(wb: Workbook) -> bytes:
    """Workbook obyektini bytes ga o'tkazish."""
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


async def export_all_excel(db: AsyncSession, owner_id: int) -> bytes:
    """Barcha guruhlar va olimpiadalar kesimida barcha natijalar."""
    stmt = (
        select(
            Group.name,
            Test.title,
            Student.full_name,
            Attempt.score,
            Attempt.total,
            Attempt.percent,
            Attempt.needs_review,
            Attempt.created_at,
        )
        .join(Titul, Titul.id == Attempt.titul_id)
        .join(Student, Student.id == Titul.student_id)
        .join(Test, Test.id == Titul.test_id)
        .join(Group, Group.id == Test.group_id)
        .where(Group.owner_id == owner_id, Attempt.status == "done")
        .order_by(Group.name, Test.title, Attempt.percent.desc().nullslast())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Barcha Natijalar"
    
    # Header
    ws.append([
        "Guruh",
        "Test / Olimpiada",
        "O'quvchi F.I.Sh.",
        "To'g'ri",
        "Jami",
        "Foiz",
        "Qayta tekshirish",
        "Sana"
    ])
    
    for r in rows:
        ws.append([
            r[0],
            r[1],
            r[2],
            r[3],
            r[4],
            float(r[5]) if r[5] is not None else 0.0,
            "Ha" if r[6] else "Yo'q",
            r[7],
        ])
        
    _style_sheet(ws)
    return _to_bytes(wb)


async def export_group_excel(db: AsyncSession, group_id: int) -> bytes:
    """Belgilangan guruhdagi barcha olimpiadalar bo'yicha natijalar."""
    stmt = (
        select(
            Test.title,
            Student.full_name,
            Attempt.score,
            Attempt.total,
            Attempt.percent,
            Attempt.needs_review,
            Attempt.created_at,
        )
        .join(Titul, Titul.id == Attempt.titul_id)
        .join(Student, Student.id == Titul.student_id)
        .join(Test, Test.id == Titul.test_id)
        .where(Test.group_id == group_id, Attempt.status == "done")
        .order_by(Test.title, Attempt.percent.desc().nullslast())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Guruh Natijalari"
    
    # Header
    ws.append([
        "Test / Olimpiada",
        "O'quvchi F.I.Sh.",
        "To'g'ri",
        "Jami",
        "Foiz",
        "Qayta tekshirish",
        "Sana"
    ])
    
    for r in rows:
        ws.append([
            r[0],
            r[1],
            r[2],
            r[3],
            float(r[4]) if r[4] is not None else 0.0,
            "Ha" if r[5] else "Yo'q",
            r[6],
        ])
        
    _style_sheet(ws)
    return _to_bytes(wb)


async def export_test_excel(db: AsyncSession, test_id: int) -> bytes:
    """Ushbu olimpiadada (testda) ishtirok etgan barcha o'quvchilar natijasi."""
    # app/services/history.py test_results() mantiqini ishlatamiz yoki select query
    stmt = (
        select(
            Student.full_name,
            Attempt.score,
            Attempt.total,
            Attempt.percent,
            Attempt.needs_review,
            Attempt.created_at,
        )
        .join(Titul, Titul.id == Attempt.titul_id)
        .join(Student, Student.id == Titul.student_id)
        .where(Titul.test_id == test_id, Attempt.status == "done")
        .order_by(Attempt.percent.desc().nullslast())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Olimpiada Natijalari"
    
    # Header
    ws.append([
        "O'quvchi F.I.Sh.",
        "To'g'ri",
        "Jami",
        "Foiz",
        "Qayta tekshirish",
        "Sana"
    ])
    
    for r in rows:
        ws.append([
            r[0],
            r[1],
            r[2],
            float(r[3]) if r[3] is not None else 0.0,
            "Ha" if r[4] else "Yo'q",
            r[5],
        ])
        
    _style_sheet(ws)
    return _to_bytes(wb)


async def export_student_excel(db: AsyncSession, student_id: int) -> bytes:
    """Ushbu o'quvchi ishtirok etgan barcha olimpiadalar/testlar."""
    stmt = (
        select(
            Test.title,
            Attempt.score,
            Attempt.total,
            Attempt.percent,
            Attempt.needs_review,
            Attempt.created_at,
        )
        .join(Titul, Titul.id == Attempt.titul_id)
        .join(Test, Test.id == Titul.test_id)
        .where(Titul.student_id == student_id, Attempt.status == "done")
        .order_by(Attempt.created_at.desc())
    )
    rows = (await db.execute(stmt)).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "O'quvchi Natijalari"
    
    # Header
    ws.append([
        "Test / Olimpiada",
        "To'g'ri",
        "Jami",
        "Foiz",
        "Qayta tekshirish",
        "Sana"
    ])
    
    for r in rows:
        ws.append([
            r[0],
            r[1],
            r[2],
            float(r[3]) if r[3] is not None else 0.0,
            "Ha" if r[4] else "Yo'q",
            r[5],
        ])
        
    _style_sheet(ws)
    return _to_bytes(wb)
